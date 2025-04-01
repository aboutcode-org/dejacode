from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def export_to_xlsx(data):
    print("Creating workbook...")  # Debug message
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan Results"

    headers = ['File', 'License', 'Vulnerabilities', 'Code Quality']
    ws.append(headers)

    for item in data:
        row = [item.get('file_name', ''),
               item.get('license', ''),
               item.get('vulnerabilities', ''),
               item.get('code_quality', '')]
        ws.append(row)

    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20

    print("Saving workbook...")  # Debug message
    try:
        wb.save('dejacode_export.xlsx')
        print("File saved successfully as dejacode_export.xlsx")
    except Exception as e:
        print(f"Error saving file: {e}")

# Example data to test the function
data = [
    {'file_name': 'file1.py', 'license': 'MIT', 'vulnerabilities': 'None', 'code_quality': 'High'},
    {'file_name': 'file2.py', 'license': 'GPL', 'vulnerabilities': 'Critical', 'code_quality': 'Low'}
]

print("Starting export...")  # Debug message
export_to_xlsx(data)
print("Export complete.")  # Debug message
